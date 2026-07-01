# -*- coding: utf-8 -*-
import csv
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

# 文件路径
csv_path = r"C:\Users\Administrator\Desktop\zhaopin\秋招公司累计汇总.csv"
excel_path = r"C:\Users\Administrator\Desktop\zhaopin\秋招公司累计汇总.xlsx"

# 读取CSV数据
data = []
with open(csv_path, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    headers = reader.fieldnames
    for row in reader:
        data.append(row)

print(f"读取到 {len(data)} 条记录")
print(f"列名: {headers}")

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "秋招公司汇总"

# 写入表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)

# 写入数据
for row_idx, row_data in enumerate(data, 2):
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

# 设置表头样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 设置列宽
column_widths = {
    "公司名称": 20,
    "公司类型": 18,
    "招聘项目/批次": 35,
    "计算机相关岗位": 50,
    "招聘链接/地址": 45,
    "截止时间": 25,
    "工作地点": 30,
    "信息来源": 22,
    "信息可信度": 12,
    "首次收录日期": 14,
    "最近更新日期": 14,
    "备注": 35,
}

for col, header in enumerate(headers, 1):
    width = column_widths.get(header, 15)
    ws.column_dimensions[get_column_letter(col)].width = width

# 设置数据行样式
data_alignment = Alignment(vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 找到信息可信度列的索引
credibility_col = None
for col, header in enumerate(headers, 1):
    if header == "信息可信度":
        credibility_col = col
        break

# 颜色定义
green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

for row in range(1, len(data) + 2):
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if row > 1:
            cell.alignment = data_alignment
    
    # 设置可信度列的颜色
    if credibility_col and row > 1:
        cell = ws.cell(row=row, column=credibility_col)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value = cell.value
        if value == "高":
            cell.fill = green_fill
        elif value == "中":
            cell.fill = yellow_fill
        elif value == "低" or value == "待核实":
            cell.fill = red_fill

# 设置行高
for row in range(1, len(data) + 2):
    ws.row_dimensions[row].height = 30

# 冻结首行
ws.freeze_panes = "A2"

# 保存文件
wb.save(excel_path)
print(f"\nExcel文件已保存到: {excel_path}")
print(f"共 {len(data)} 条记录")
