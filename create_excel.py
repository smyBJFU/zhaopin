import os
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("openpyxl not available, trying to install...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Excel文件路径
excel_path = r"C:\Users\Administrator\Desktop\zhaopin\秋招公司累计汇总.xlsx"

# 今天的日期
today = "2026-07-01"

# 公司数据
companies = [
    # 互联网大厂
    {
        "公司名称": "字节跳动",
        "公司类型": "互联网大厂",
        "招聘项目/批次": "Seed大模型人才校招/前沿技术领域人才校招（提前批）",
        "计算机相关岗位": "大模型算法、多模态算法、智能体研发、后端开发、前端开发、客户端开发、算法工程师、数据研发",
        "招聘链接/地址": "https://jobs.bytedance.com/campus",
        "截止时间": "招满即止",
        "工作地点": "北京、上海、深圳、杭州、新加坡、圣何塞、西雅图",
        "信息来源": "字节跳动官网、牛客网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "阿里巴巴",
        "公司类型": "互联网大厂",
        "招聘项目/批次": "阿里星顶尖人才计划（提前批）",
        "计算机相关岗位": "基础模型算法、AI Infra、大模型应用、产业AI、计算架构、芯片研发、后端开发、前端开发",
        "招聘链接/地址": "https://campus-talent.alibaba.com/activity/ali-star?lang=zh",
        "截止时间": "待公布",
        "工作地点": "杭州、北京、上海、深圳",
        "信息来源": "阿里巴巴校招官网、高校就业网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "腾讯",
        "公司类型": "互联网大厂",
        "招聘项目/批次": "青云计划（提前批）",
        "计算机相关岗位": "大语言模型/NLP、多模态/视觉/语音、智能体、强化学习、搜索推荐、AI基础设施、系统与安全、AI数据工程",
        "招聘链接/地址": "https://join.qq.com/qingyun.html",
        "截止时间": "待公布",
        "工作地点": "深圳、北京、上海",
        "信息来源": "腾讯校招官网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "百度",
        "公司类型": "互联网大厂",
        "招聘项目/批次": "AIDU计划（提前批）",
        "计算机相关岗位": "大模型算法、多模态算法、智能体、自动驾驶、AI异构计算、语音技术、后端开发、前端开发",
        "招聘链接/地址": "https://talent.baidu.com/jobs/list",
        "截止时间": "待公布",
        "工作地点": "北京、深圳",
        "信息来源": "新华网、百度官网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "拼多多",
        "公司类型": "互联网大厂",
        "招聘项目/批次": "云弧计划/研发实习生（提前批）",
        "计算机相关岗位": "大模型算法、AI Infra、服务端研发、Web前端、客户端研发、算法、网络安全",
        "招聘链接/地址": "https://careers.pddglobalhr.com/campus",
        "截止时间": "2026-08-31",
        "工作地点": "上海",
        "信息来源": "牛客网、超级简历",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "京东",
        "公司类型": "互联网大厂",
        "招聘项目/批次": "TGT顶尖青年技术天才计划（提前批）",
        "计算机相关岗位": "多模态大模型、空间与具身智能、机器学习、AI Infra、搜索推荐广告、高性能与云计算、安全、大数据",
        "招聘链接/地址": "https://campus.jd.com",
        "截止时间": "待公布",
        "工作地点": "北京",
        "信息来源": "牛客网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "蚂蚁集团",
        "公司类型": "金融科技",
        "招聘项目/批次": "蚂蚁星PlanA顶尖人才专项（提前批）",
        "计算机相关岗位": "大模型Agent工程、持续学习算法、AIGC安全、大模型后训练、金融AI",
        "招聘链接/地址": "蚂蚁集团官网",
        "截止时间": "2026-07-18",
        "工作地点": "杭州、北京、上海、福建",
        "信息来源": "超级简历",
        "首次收录日期": today,
        "最近更新日期": today
    },
    # 央国企
    {
        "公司名称": "国家电网",
        "公司类型": "央国企",
        "招聘项目/批次": "提前批/第一批",
        "计算机相关岗位": "计算机科学与技术、软件工程、通信工程、自动化、电力系统及其自动化",
        "招聘链接/地址": "https://zhaopin.sgcc.com.cn",
        "截止时间": "提前批：2026年7-9月；一批：预计11月中旬",
        "工作地点": "全国各省市",
        "信息来源": "中公教育、思格教育",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国移动",
        "公司类型": "央国企",
        "招聘项目/批次": "技术岗提前批",
        "计算机相关岗位": "软件开发、系统开发、网络技术、大数据开发、云计算、信息安全",
        "招聘链接/地址": "https://www.iguopin.com（国聘网）",
        "截止时间": "预计7月底开放",
        "工作地点": "全国各省市",
        "信息来源": "高顿教育、国聘网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国联通",
        "公司类型": "央国企",
        "招聘项目/批次": "技术岗提前批",
        "计算机相关岗位": "软件开发、系统开发、网络技术、大数据、云计算",
        "招聘链接/地址": "https://www.iguopin.com（国聘网）",
        "截止时间": "预计7月底开放",
        "工作地点": "全国各省市",
        "信息来源": "高顿教育、国聘网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国电信",
        "公司类型": "央国企",
        "招聘项目/批次": "技术岗提前批",
        "计算机相关岗位": "软件开发、系统开发、网络技术、大数据、云计算",
        "招聘链接/地址": "https://www.iguopin.com（国聘网）",
        "截止时间": "预计7月底开放",
        "工作地点": "全国各省市",
        "信息来源": "高顿教育、国聘网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国航天科技集团",
        "公司类型": "央国企",
        "招聘项目/批次": "提前批",
        "计算机相关岗位": "计算机科学与技术、软件工程、人工智能、控制科学与工程、信息与通信工程、航天算法",
        "招聘链接/地址": "各院所官网/国聘网",
        "截止时间": "预计8月发布",
        "工作地点": "北京、上海、西安、成都、武汉",
        "信息来源": "今日头条",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国航天科工集团",
        "公司类型": "央国企",
        "招聘项目/批次": "提前批",
        "计算机相关岗位": "计算机科学与技术、软件工程、人工智能、电子科学与技术、测控技术",
        "招聘链接/地址": "各院所官网/国聘网",
        "截止时间": "预计8月发布",
        "工作地点": "北京、上海、西安、成都、武汉",
        "信息来源": "今日头条",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国石油（中石油）",
        "公司类型": "央国企",
        "招聘项目/批次": "提前批/暑期定向招录",
        "计算机相关岗位": "信息技术岗、软件开发、数据分析、网络安全、自动化",
        "招聘链接/地址": "中石油招聘官网/国聘网",
        "截止时间": "预计8月发布",
        "工作地点": "全国各省市",
        "信息来源": "高顿教育",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国石化（中石化）",
        "公司类型": "央国企",
        "招聘项目/批次": "提前批/暑期定向招录",
        "计算机相关岗位": "信息技术岗、软件开发、数据分析、网络安全",
        "招聘链接/地址": "中石化招聘官网/国聘网",
        "截止时间": "预计8月发布",
        "工作地点": "全国各省市",
        "信息来源": "高顿教育",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国储备粮管理集团（中储粮）",
        "公司类型": "央国企",
        "招聘项目/批次": "2027届秋招批次",
        "计算机相关岗位": "信息技术岗",
        "招聘链接/地址": "https://sinograin2026.iguopin.com",
        "截止时间": "2026年6月30日（已截止，供参考）",
        "工作地点": "全国各省分公司",
        "信息来源": "国聘网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中兴通讯",
        "公司类型": "通信科技/央国企",
        "招聘项目/批次": "2027届秋招",
        "计算机相关岗位": "软件开发、人工智能算法、微电子、光电子、计算机科学与技术",
        "招聘链接/地址": "中兴通讯招聘官网",
        "截止时间": "招满为止",
        "工作地点": "深圳、广州",
        "信息来源": "高校就业网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    # 银行
    {
        "公司名称": "中国工商银行",
        "公司类型": "金融科技/银行",
        "招聘项目/批次": "秋季校园招聘",
        "计算机相关岗位": "信息科技岗、软件开发、数据分析、网络安全、金融科技岗",
        "招聘链接/地址": "工行招聘官网",
        "截止时间": "预计9月5日-10月9日",
        "工作地点": "全国各省市",
        "信息来源": "新浪财经、搜狐网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国农业银行",
        "公司类型": "金融科技/银行",
        "招聘项目/批次": "秋季校园招聘",
        "计算机相关岗位": "信息科技岗、软件开发、数据分析、网络安全",
        "招聘链接/地址": "农行招聘官网",
        "截止时间": "预计9月5日-10月9日",
        "工作地点": "全国各省市",
        "信息来源": "新浪财经、搜狐网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国银行",
        "公司类型": "金融科技/银行",
        "招聘项目/批次": "秋季校园招聘",
        "计算机相关岗位": "信息科技岗、软件开发、数据分析、金融科技",
        "招聘链接/地址": "中行招聘官网",
        "截止时间": "预计9月5日-10月10日",
        "工作地点": "全国各省市",
        "信息来源": "新浪财经、搜狐网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "中国建设银行",
        "公司类型": "金融科技/银行",
        "招聘项目/批次": "秋季校园招聘",
        "计算机相关岗位": "信息科技岗、软件开发、数据分析、网络安全",
        "招聘链接/地址": "建行招聘官网",
        "截止时间": "预计9月9日-10月10日",
        "工作地点": "全国各省市",
        "信息来源": "新浪财经、搜狐网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    # 硬科技
    {
        "公司名称": "大疆创新",
        "公司类型": "智能制造/硬科技",
        "招聘项目/批次": "2027届校招",
        "计算机相关岗位": "嵌入式开发、算法工程师、软件开发、智能硬件研发",
        "招聘链接/地址": "大疆招聘官网",
        "截止时间": "待公布",
        "工作地点": "深圳",
        "信息来源": "官方微博/校招平台",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "科大讯飞",
        "公司类型": "智能制造/硬科技",
        "招聘项目/批次": "2027届校招",
        "计算机相关岗位": "语音算法、NLP算法、大模型研发、软件开发、测试开发",
        "招聘链接/地址": "科大讯飞招聘官网",
        "截止时间": "待公布",
        "工作地点": "合肥、北京、上海、深圳",
        "信息来源": "力扣、高校就业网",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "TP-LINK",
        "公司类型": "智能制造/硬科技",
        "招聘项目/批次": "提前批",
        "计算机相关岗位": "嵌入式开发、网络协议开发、软件测试、硬件开发",
        "招聘链接/地址": "TP-LINK招聘官网",
        "截止时间": "招满即止",
        "工作地点": "深圳、杭州",
        "信息来源": "今日头条",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "长鑫存储",
        "公司类型": "智能制造/硬科技",
        "招聘项目/批次": "提前批",
        "计算机相关岗位": "芯片设计、验证工程师、软件开发、算法工程师",
        "招聘链接/地址": "长鑫存储招聘官网",
        "截止时间": "招满即止",
        "工作地点": "合肥、上海、北京",
        "信息来源": "今日头条",
        "首次收录日期": today,
        "最近更新日期": today
    },
    {
        "公司名称": "长亭科技",
        "公司类型": "网络安全/硬科技",
        "招聘项目/批次": "安全服务校招提前批",
        "计算机相关岗位": "网络安全工程师、安全研究员、渗透测试",
        "招聘链接/地址": "长亭科技招聘官网",
        "截止时间": "待公布",
        "工作地点": "北京、上海、杭州",
        "信息来源": "超级简历",
        "首次收录日期": today,
        "最近更新日期": today
    },
]

# 检查文件是否存在
if os.path.exists(excel_path):
    print(f"Excel文件已存在，读取并更新...")
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 读取现有数据，构建字典（公司名称+批次为key）
    existing_data = {}
    for row in range(2, ws.max_row + 1):
        company_name = ws.cell(row=row, column=1).value
        batch = ws.cell(row=row, column=3).value
        key = f"{company_name}_{batch}"
        existing_data[key] = row
    
    print(f"现有记录数: {len(existing_data)}")
    
    # 更新或添加新数据
    new_count = 0
    update_count = 0
    for company in companies:
        key = f"{company['公司名称']}_{company['招聘项目/批次']}"
        if key in existing_data:
            # 更新已有记录
            row_num = existing_data[key]
            # 只更新最近更新日期和可能变化的字段
            ws.cell(row=row_num, column=9).value = company["最近更新日期"]
            update_count += 1
        else:
            # 添加新记录
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1).value = company["公司名称"]
            ws.cell(row=row_num, column=2).value = company["公司类型"]
            ws.cell(row=row_num, column=3).value = company["招聘项目/批次"]
            ws.cell(row=row_num, column=4).value = company["计算机相关岗位"]
            ws.cell(row=row_num, column=5).value = company["招聘链接/地址"]
            ws.cell(row=row_num, column=6).value = company["截止时间"]
            ws.cell(row=row_num, column=7).value = company["工作地点"]
            ws.cell(row=row_num, column=8).value = company["信息来源"]
            ws.cell(row=row_num, column=9).value = company["首次收录日期"]
            ws.cell(row=row_num, column=10).value = company["最近更新日期"]
            new_count += 1
    
    print(f"新增记录: {new_count}条")
    print(f"更新记录: {update_count}条")
    print(f"总记录数: {ws.max_row - 1}条")
    
else:
    print(f"创建新的Excel文件...")
    wb = Workbook()
    ws = wb.active
    ws.title = "秋招公司汇总"
    
    # 设置表头
    headers = ["公司名称", "公司类型", "招聘项目/批次", "计算机相关岗位", "招聘链接/地址", 
               "截止时间", "工作地点", "信息来源", "首次收录日期", "最近更新日期"]
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 写入数据
    for row, company in enumerate(companies, 2):
        ws.cell(row=row, column=1).value = company["公司名称"]
        ws.cell(row=row, column=2).value = company["公司类型"]
        ws.cell(row=row, column=3).value = company["招聘项目/批次"]
        ws.cell(row=row, column=4).value = company["计算机相关岗位"]
        ws.cell(row=row, column=5).value = company["招聘链接/地址"]
        ws.cell(row=row, column=6).value = company["截止时间"]
        ws.cell(row=row, column=7).value = company["工作地点"]
        ws.cell(row=row, column=8).value = company["信息来源"]
        ws.cell(row=row, column=9).value = company["首次收录日期"]
        ws.cell(row=row, column=10).value = company["最近更新日期"]
    
    print(f"新增记录: {len(companies)}条")
    print(f"总记录数: {len(companies)}条")

# 设置列宽
column_widths = [18, 16, 35, 50, 40, 25, 30, 20, 15, 15]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

# 设置行高和对齐
for row in range(1, ws.max_row + 1):
    ws.row_dimensions[row].height = 30
    for col in range(1, 11):
        cell = ws.cell(row=row, column=col)
        if row > 1:  # 数据行
            cell.alignment = Alignment(vertical="center", wrap_text=True)

# 添加边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in range(1, ws.max_row + 1):
    for col in range(1, 11):
        ws.cell(row=row, column=col).border = thin_border

# 冻结首行
ws.freeze_panes = "A2"

# 保存文件
wb.save(excel_path)
print(f"\nExcel文件已保存到: {excel_path}")
